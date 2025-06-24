import time
import win32com.client
import os
import re
from openpyxl import load_workbook

while True:
    try:
        # PLACE ALL YOUR EXISTING CODE HERE
        EXCEL_FILE = "test2.xlsx"  # Your Excel sheet
        SHEET_NAME = "Sheet1"  # Adjust if your sheet has a custom name

        # Load workbook and sheet
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]

        # Build mapping from Employee ID to row index
        id_to_row = {}
        for row in range(2, ws.max_row + 1):
            emp_id = str(ws.cell(row=row, column=2).value)  # employee ID in column 2
            id_to_row[emp_id] = row

        # Define document columns
        doc_columns = {
            "resume": 4,
            "salary slip": 5,
            "aadhar card": 6,
            "pan card": 7
        }

        # Outlook setup
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)  # Inbox
        messages = inbox.Items
        messages.Sort("[ReceivedTime]", True)

        # Scan emails
        for msg in list(messages)[:5]:
            try:
                if msg.Class != 43:  # MailItem only
                    continue

                content = msg.Subject + " " + msg.Body
                id_match = re.search(r'\bID[:\- ]+(\d+)', content, re.IGNORECASE)
                doc_match = re.search(r'\b(document[:\- ]*)?(resume|salary slip|aadhar card|pan card)', content, re.IGNORECASE)

                if not id_match or not doc_match:
                    print(f"Skipping email: Missing ID or document type → {msg.Subject}")
                    continue

                emp_id = id_match.group(1).strip()
                doc_type = doc_match.group(2).lower().strip()

                if emp_id not in id_to_row:
                    print(f"❌ ID not found in sheet: {emp_id}")
                    continue

                if doc_type not in doc_columns:
                    print(f"❌ Unknown document type: {doc_type}")
                    continue

                row = id_to_row[emp_id]
                col = doc_columns[doc_type]

                # Check if at least one attachment exists
                if msg.Attachments.Count > 0:
                    ws.cell(row=row, column=col).value = "Approved"
                    print(f"✅ Updated: ID {emp_id} | {doc_type.title()} → Approved")
                else:
                    print(f"⚠️ Attachment missing: ID {emp_id} | {doc_type.title()} → Skipped")


            except Exception as e:
                print(f"⚠️ Error: {e}")

        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"\n📘 Excel updated: {EXCEL_FILE}")



    except Exception as e:
        print(f"Unexpected error: {e}")

    time.sleep(30)  # wait 30 seconds before next run
